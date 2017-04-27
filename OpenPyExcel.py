import openpyxl
import re
from bs4 import BeautifulSoup
import urllib.request
import xlsxwriter

#INPUT SHEET INFO
wb = openpyxl.load_workbook('InputSet.xlsx')
sheet = wb.get_sheet_by_name('UrlBook')
r = sheet.max_row
c = sheet.max_column

# OUTPUT SHEET INFO
workbook = xlsxwriter.Workbook('Analysis.xlsx')
heading = workbook.add_format({'bold':True,'font_color':'red'})
chart1 = workbook.add_chart({'type':'column'})


column_let = openpyxl.cell.cell.get_column_letter(c)
column_let = column_let + str(r)
print(column_let)
#print(openpyxl.cell.cell.get_column_letter(c))
row_data = []
for rowOfCellObjects in sheet['A1':column_let]:
    for cellObj in rowOfCellObjects:
        #print(cellObj.coordinate, cellObj.value)
        row_data.append(cellObj.value)

    #excel file data is extracted
    #performing webScrapping with BeautifulSoup
    req = urllib.request.Request(row_data[0], data = None, headers = {'User-Agent':' Moziila/5.0 (Macintosh;Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari / 537.36'})
    f = urllib.request.urlopen(req)
    s = f.read().decode('utf-8')
    soup = BeautifulSoup(s,"html.parser")
    for script in soup(["script","style"]):
        script.extract()

    #extracting text information
    text = soup.get_text()
    print("total character ",len(text))
    print(row_data)

    #adding worksheet with table headings
    worksheet = workbook.add_worksheet()
    worksheet.write("A1","Kewords",heading)
    worksheet.write("B1","No of Occurence",heading)
    worksheet.write("C1","Weightage",heading)
 

    #processing keywords.....
    keyword1 = re.findall(row_data[1],text)
    worksheet.write("A2",row_data[1])
    worksheet.write("B2",len(keyword1))
    worksheet.write("C2",(len(keyword1)/len(text))*100)
    
    
    keyword2 = re.findall(row_data[2],text)
    worksheet.write("A3",row_data[2])
    worksheet.write("B3",len(keyword2))
    worksheet.write("C3",(len(keyword2)/len(text))*100)
    
    
    keyword3 = re.findall(row_data[3],text)
    worksheet.write("A4",row_data[3])
    worksheet.write("B4",len(keyword3))
    worksheet.write("C4",(len(keyword3)/len(text))*100)

    
    keyword4 = re.findall(row_data[4],text)
    worksheet.write("A5",row_data[4])
    worksheet.write("B5",len(keyword4))
    worksheet.write("C5",(len(keyword4)/len(text))*100)
    
    #calculating remaing words on Page
    worksheet.write("A6","Other Words on page",heading)
    oth = len(text)-len(keyword1)-len(keyword2)-len(keyword3)-len(keyword4)
    worksheet.write("B6",oth,heading)
    worksheet.write("C6",(oth/len(text))*100,heading)

    #retriving Worksheet Name to create Formula
    s_name = worksheet.name
    formula = '='+s_name+'!$C2:$C6'
    print(formula)
    #adding Chart to Worksheet
    chart1.add_series({'values':formula})
    worksheet.insert_chart("A8",chart1)

    #resetting row data here to process next row of information
    row_data = []

workbook.close()
